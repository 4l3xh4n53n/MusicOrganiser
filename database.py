"""
This file handles the database. The database used is a xlsx file.

Manipulating the spreadsheet is proving to be a rather time consuming task to accomplish.
I think an easier solution would be to read the spreadsheet line by line and turn it into
a database instead and use a database from here on out.
A database will be easier to interact with, faster, more stable, among many other things.
I also don't think I will interact with the spreadsheet much if the application has
features that allow me to search for incomplete albums.

"""
from openpyxl.worksheet.worksheet import Worksheet

# TODO
"""
when adding an artists along with albums, or adding multiple types of albums make it 
find the position of the artist and hold it, so it doesn't need to be found each time 
we want to get add a new album (which may happen multiple times when adding a new artist
"""

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, PatternFill

"""
Markers:
I might change these to priorities!

Markers for artists:
G = get something
F = finish discography

Markers for albums:
G = get this album
"""

# This function returns the albums for an artist
def get_artist(artist_name):
    workbook = load_workbook("test.xlsx")
    sheet = workbook.active

    sheet["A1"] = "|"
    sheet["B1"] = "||"
    sheet["A2"] = "||"
    sheet["B2"] = "|_"

    sheet.insert_rows(0, 2)

    workbook.save(filename="test.xlsx")

    print("Getting Artist!")

"""
This function adds different headers
Artist headers span cells A:I and are #B4C7DC
Album type headers span cells A:C and are #FF972F
"""
def add_header(type): # TODO, I don't know how to work this yet (might just not bother with this)
    if type == "artist":
        print("Adding artist")
    print("Adding a new header")

"""
This function adds new rows to the spreadsheet while keeping the formatting 
of everything under the new rows. Care should be taken not to leave rows blank
otherwise search functions may end prematurely.
"""
def add_new_rows(sheet, position, number_of_rows):

    sheet.insert_rows(position, number_of_rows)

    # unmerge and remerge innit

    # Create a copy to stop iteration error
    merged_cells = list(sheet.merged_cells)

    # Un-merge cells

    for merged_cell_range in merged_cells:
        merged_row = merged_cell_range.min_row
        if merged_row > position:
            # For some reason which I cannot figure out, sheet.unmerge_cells always complains it
            # cannot find the cells to unmerge and the code "fails" however, the cells are still un-merged
            #  without this code the cells remain merged
            try:
                sheet.unmerge_cells(range_string="A" + str(merged_row) + ":I" + str(merged_row))
            except KeyError:
                pass

    # Re-merge cells

    for merged_cell_range in merged_cells:
        merged_row = merged_cell_range.min_row
        if merged_row >= position:
            new_cell_row = str(merged_cell_range.min_row + number_of_rows)

            # todo THIS DOESN'T WORK
            # Cell is an artist cell
            if sheet["A" + str(merged_row)].fill == PatternFill("solid", fgColor="B4C7DC"):
                sheet.merge_cells("A" + new_cell_row + ":I" + new_cell_row)

            # Cell is an album type cell
            else:
                sheet.merge_cells("A" + new_cell_row + ":C" + new_cell_row)

    return sheet # TODO, could sheet be put in the outer scope of this file?


"""
This function adds a new artist to the spreadsheet. All albums under the name of 
the artist will be stored under the artist cell.
This function should only be run after making sure the artist doesn't already exist. 
"""
def add_artist(artist:str):
    workbook = load_workbook("test.xlsx")
    sheet = workbook.active

    # Find out where in the spreadsheet the artist should be positioned

    previous_artist_position = 1

    for row_index, row in enumerate(sheet.iter_rows(min_row=1), start=1): # todo, optimise by looping through merged cells

        # If the end of the spreadsheet has been reached
        if row[0].value is None:
            previous_artist_position = row_index
            break

        # Figure out if the cell contains an artists name by its background colour
        if row[0].fill == PatternFill("solid", fgColor="B4C7DC"):

            # The new artist (alphabetically) comes before this artist (found position and finish loop)
            if artist < row[0].value:
                previous_artist_position = row_index
                break

            # The new artist (alphabetically) comes after this artist (continue loop)
            else:
                previous_artist_position = row_index

    add_new_rows(sheet, previous_artist_position, 1)

    new_artist_position = str(previous_artist_position)

    # Add the title
    sheet["A" + new_artist_position] = artist

    # Change the formatting
    sheet.merge_cells("A" + new_artist_position + ":I" + new_artist_position)
    sheet["A" + new_artist_position].alignment = Alignment(horizontal="center")
    sheet["A" + new_artist_position].fill = PatternFill("solid", fgColor="B4C7DC")

    workbook.save("test.xlsx")

"""
This function finds an artist and returns it's position
"""
def find_artist(sheet:Worksheet, artist):

    # Loop all merged cells
    # make sure merged cell has correct background colour
    # make sure it's the right album

    merged_cells = sheet.merged_cells

    for merged_cell in merged_cells:
        cell = merged_cell.min_row[0]
        if cell.fill == PatternFill("solid", fgColor="B4C7DC"):
            if cell.value == artist:
                return cell.row

    return None

"""
This function adds new albums under a certain artist
The type value passed to this function can be any string value however should be a type of albums such as:
- Studio Album
- Extended Play
- Compilation Album ...
"""
def add_albums(artist, type, albums):

    # Find artist (probably also a function)
    # Insert new blank lines (this is a function)
    # Add the type of album header if it doesn't already exist
    # Make sure it's ordered correctly (by date)

    print("Adding album of type: " + type)

"""
This function adds an artist column to the spreadsheet.
Currently it can also add albums but this functionality will be moved
"""
def add_artist_old(artist, albums):

    def insert_empty_rows():
        sheet.insert_rows(previous_artist_position, 1 + len(albums))

    """
    This function un-merges and re-merges cells in order to keep the formatting of the 
    spreadsheet when new rows are inserted for storing new artists and/or albums.
    
    For reference:
    - Artist cells, span columns A to I and are the name of the artist all albums are found under
    their background colour is: #B4C7DC
    - Album type cells, span columns A to C and differentiates different types of albums, such as 
    compilation, EP, Studio, etc... 
    their background colour is: #FF972F
    """
    def format_cells():

        # TODO, add formatting for headings like singles and compilations as these have separate rules
        # todo this I would have to add albums, singles, compilations, ep's as multiple separate thingies (I may handle this elsewhere in the code)

        # Create a copy to stop iteration error
        merged_cells = list(sheet.merged_cells)

        # Un-merge cells

        for merged_cell_range in merged_cells:
            merged_row = merged_cell_range.min_row
            if merged_row > new_artist_position:
                # For some reason which I cannot figure out, sheet.unmerge_cells always complains it
                # cannot find the cells to unmerge and the code "fails" however, the cells are still un-merged
                #  without this code the cells remain merged
                try:
                    sheet.unmerge_cells(range_string="A" + str(merged_row) + ":I" + str(merged_row))
                except KeyError:
                    pass

        # Re-merge cells

        for merged_cell_range in merged_cells:
            merged_row = merged_cell_range.min_row
            if merged_row >= new_artist_position:
                new_cell_row = str(merged_cell_range.min_row + 1 + len(albums))

                # Cell is an artist cell
                if row[0].fill == PatternFill("solid", fgColor="B4C7DC"):
                    sheet.merge_cells("A" + new_cell_row + ":I" + new_cell_row)

                # Cell is an album type cell
                else:
                    sheet.merge_cells("A" + new_cell_row + ":C" + new_cell_row)

    workbook = load_workbook("test.xlsx")
    sheet = workbook.active

    # Find out where in the spreadsheet the artist should be positioned

    previous_artist_position = 1

    for row_index, row in enumerate(sheet.iter_rows(min_row=1), start=1):

        # If the end of the spreadsheet has been reached
        if row[0].value is None:
            previous_artist_position = row_index
            insert_empty_rows()
            break

        # Figure out if the cell contains an artists name by its background colour
        if row[0].fill == PatternFill("solid", fgColor="B4C7DC"):

            # The new artist (alphabetically) comes before this artist (found position and finish loop)
            if artist < row[0].value:
                previous_artist_position = row_index
                insert_empty_rows()
                break

            # The new artist (alphabetically) comes after this artist (continue loop)
            else:
                previous_artist_position = row_index

    new_artist_position = previous_artist_position

    format_cells()

    # Merges new cell and sets content and formatting
    sheet.merge_cells("A" + str(new_artist_position) + ":I" + str(new_artist_position))
    sheet["A" + str(new_artist_position)] = artist
    sheet["A" + str(new_artist_position)].alignment = Alignment(horizontal="center")
    sheet["A" + str(new_artist_position)].fill = PatternFill("solid", fgColor="B4C7DC")

    # Add each of the albums under the artist

    for album_index, album in enumerate(albums, start=1):
        album_position = str(new_artist_position + album_index)
        sheet["A" + album_position] = album[0]
        sheet["B" + album_position] = album[1]

    workbook.save("test.xlsx")
