"""
This file handles everything that should migrate my current spreadsheet over to mongoDB.
This will likely only be used once and never again.
"""
import json

from openpyxl import load_workbook
from pymongo import MongoClient

def migrate():

    print("Migrating to database.")

    # Create an empty "JSON" object
    music_data = []

    # Open the sheet
    workbook = load_workbook("./Checklist.xlsx")
    sheet = workbook.active

    current_artist = None
    current_album_type = ""

    # Read the spreadsheet one file at a time
    for row_index, row in enumerate(sheet.iter_rows(min_row=1), start=1):

        cell1 = row[0]
        cell_color = cell1.fill.start_color.index

        if cell_color == "00000000":
            print("Found the end of the file")
            break

        # If cell contains artist name
        elif cell_color == "FFB4C7DC":

            # Append the previous artist
            if current_artist is not None:
                music_data.append(current_artist)

            artist = cell1.value
            # Studio albums come first in the spreadsheet and have no header
            current_album_type = "studio_album"

            # Set blank template to be used later

            current_artist = {
                "name": artist,
                "albums": [],
                "markers": row[9].value,
                "notes": row[10].value
            }

        # If cell contains album type
        elif cell_color == "FFFF972F":
            current_album_type = cell1.value.lower()

            # Convert names to be singular before being stored in the JSON
            match current_album_type:
                case "eps":
                    current_album_type = "ep"
                case "compilations":
                    current_album_type = "compilation_album"
                case "covers":
                    current_album_type = "cover_album"
                case "singles":
                    current_album_type = "single"


        # If cell contains an album
        elif cell_color == "FFDEE6EF":

            album_json = {
                "title": cell1.value,
                "type": current_album_type,
                "date": row[1].value,
                "downloading": row[2].value.lower() == "y",
                "downloaded": row[3].value.lower() == "y", # This isn't integrated yet
                "tags": row[4].value.lower() == "y",
                "cover": row[5].value.lower() == "y",
                "replay_gain": row[6].value.lower() == "y",
                "server_upload": row[7].value.lower() == "y",
                "format": row[8].value,
                "markers": row[9].value,
                "notes": row[10].value
            }

            # append to list of current album type!
            current_artist["albums"].append(album_json)

        # If an unexpected cell colour arrives
        else:
            print("Cell color: " + cell_color)
            print("Something has gone wrong")


    print("Finished JSON conversion")
    print(json.dumps(music_data, indent=2))

    # Upload to the database

    client = MongoClient("mongodb://test:test@localhost:27017/")
    db = client.music.music
    db.create_index("name", unique=True) # Allows for quick search of the name field

    try:
        db.insert_many(music_data)
        print("Uploaded to database")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        client.close()


migrate()